from app.models import User, Enrollment, Activity, Event, Course, db
from app.repositories.user_repository import UserRepository
from app.services.event_service import EventService
from app.services.notification_service import NotificationService
from sqlalchemy import or_, func
from flask import current_app
from datetime import datetime
import csv
import io
from openpyxl import load_workbook
from app.utils import normalize_cpf

class AdminService:
    """Service layer for administrative user management tasks.
    
    Handles user listing, creation, updates, and deletion with 
    advanced filtering and pagination support.
    """
    def __init__(self):
        self.user_repo = UserRepository()
        self.event_service = EventService()
        self.notification_service = NotificationService()

    @staticmethod
    def _coerce_bool(value):
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if value is None:
            return False
        return str(value).strip().lower() in {'1', 'true', 't', 'sim', 'yes', 'y', 'on'}

    @staticmethod
    def _normalize_import_header(header_name):
        normalized = str(header_name or '').strip().lstrip('\ufeff').lower()
        aliases = {
            'curso_nome': 'curso',
            'perfil': 'role',
            'senha': 'password',
            'nome_completo': 'nome',
            'permitir_criar_eventos': 'can_create_events',
            'permissao_criar_eventos': 'can_create_events',
            'pode_criar_eventos': 'can_create_events',
        }
        return aliases.get(normalized, normalized)

    @staticmethod
    def _normalize_import_role(value):
        normalized = str(value or '').strip().lower()
        aliases = {
            'admin': 'admin',
            'professor': 'professor',
            'prof': 'professor',
            'coordenador': 'coordenador',
            'coordenacao': 'coordenador',
            'coordenação': 'coordenador',
            'coord': 'coordenador',
            'gestor': 'gestor',
            'extensao': 'extensao',
            'extensão': 'extensao',
            'participante': 'participante',
            'aluno': 'participante',
            'discente': 'participante',
        }
        return aliases.get(normalized, normalized)

    def list_users_paginated(self, page=1, per_page=10, filters=None):
        """Retrieves a paginated list of users with optional filtering.
        
        Filters can include: ra, curso, cpf, email, event_id, activity_id.
        """
        query = User.query
        
        if filters:
            if filters.get('ra'):
                query = query.filter(User.ra.ilike(f"%{filters['ra']}%"))
            if filters.get('curso'):
                query = query.join(Course, User.course_id == Course.id, isouter=True)
                query = query.filter(Course.nome.ilike(f"%{filters['curso']}%"))
            if filters.get('cpf'):
                cpf_digits = normalize_cpf(filters['cpf'])
                if cpf_digits:
                    query = query.filter(User.cpf.ilike(f"%{cpf_digits}%"))
            if filters.get('email'):
                query = query.filter(User.email.ilike(f"%{filters['email']}%"))
            if filters.get('nome'):
                query = query.filter(User.nome.ilike(f"%{filters['nome']}%"))
            
            # Filter by Event or Activity (requires joins)
            if filters.get('event_id') or filters.get('activity_id'):
                query = query.join(Enrollment, User.cpf == Enrollment.user_cpf)
                if filters.get('event_id'):
                    query = query.join(Activity, Enrollment.activity_id == Activity.id)
                    query = query.filter(Activity.event_id == filters['event_id'])
                if filters.get('activity_id'):
                    query = query.filter(Enrollment.activity_id == filters['activity_id'])

        return query.paginate(page=page, per_page=per_page, error_out=False)

    def buscar_usuarios_inscricao(self, termo):
        """Searches for users by normalized name, RA, or CPF."""
        from app.utils import normalizar_texto
        
        termo_norm = normalizar_texto(termo)
        if not termo_norm: return []
        
        all_users = User.query.all()
        matches = []
        
        for u in all_users:
            # Check RA or CPF (exact normalized match or contains)
            if termo_norm in normalizar_texto(u.ra) or termo_norm in normalizar_texto(u.cpf):
                matches.append(u)
                continue
            
            # Check Name (contains)
            if termo_norm in normalizar_texto(u.nome):
                matches.append(u)
        
        return matches[:10] # Limit to 10 results for performance

    def create_user(self, data):
        """Creates a new user manually."""
        cpf = normalize_cpf(data.get('cpf'))
        if not cpf or len(cpf) != 11:
            return None, "CPF inválido. Informe 11 dígitos."

        username = cpf

        if db.session.get(User, username) or User.query.filter_by(cpf=cpf).first():
            return None, "Usuário ou CPF já cadastrado."
            
        user = User(
            username=username,
            nome=data.get('nome'),
            email=data.get('email'),
            cpf=cpf,
            ra=data.get('ra'),
            curso=data.get('curso'),
            role=data.get('role', 'participante'),
            can_create_events=self._coerce_bool(data.get('can_create_events')),
        )
        user.set_password(data.get('password') or cpf)
        db.session.add(user)
        db.session.commit()
        return user, "Usuário criado com sucesso."

    def update_user_details(self, target_username: str, data: dict):
        """Updates an existing user's profile details."""
        user = self.user_repo.get_by_username(target_username)
        if not user:
            return False
            
        user.nome = data.get('nome', user.nome)
        user.cpf = data.get('cpf', user.cpf)
        user.email = data.get('email', user.email)
        user.ra = data.get('ra', user.ra)
        user.curso = data.get('curso', user.curso)
        user.role = data.get('role', user.role)
        if 'can_create_events' in data:
            user.can_create_events = self._coerce_bool(data.get('can_create_events'))
        
        if data.get('password'):
            user.set_password(data['password'])
            
        self.user_repo.update()
        return True

    def delete_user(self, username):
        """Deletes a user and their associations."""
        user = db.session.get(User, username)
        if not user: return False
        db.session.delete(user)
        db.session.commit()
        return True

    def manual_enroll(self, user_cpf, activity_id, actor_user=None, category_id=None):
        """Manually enrolls a user into an activity."""
        try:
            activity_id = int(activity_id)
        except (TypeError, ValueError):
            return False, "Atividade inválida."

        user = self.user_repo.get_by_cpf(user_cpf)
        activity = db.session.get(Activity, activity_id)
        
        if not user or not activity:
            return False, "Usuário ou Atividade não encontrados."
            
        success, message, enrollment = self.event_service.manual_enroll_user(
            actor_user or user,
            user,
            activity_id,
            category_id=category_id,
        )
        if not success:
            return False, message

        try:
            self._notify_manual_enrollment(user, activity)
        except Exception:
            current_app.logger.exception(
                "Falha ao enfileirar email de inscricao manual para o usuario %s na atividade %s",
                user.username,
                activity_id,
            )
        return True, message

    def _notify_manual_enrollment(self, user, activity):
        """Notify a participant when staff manually add them to an event activity."""
        if not user or not user.email or not activity:
            return

        event = activity.event
        app_url = (current_app.config.get('BASE_URL') or '').rstrip('/')
        event_path = f"/inscrever/{event.token_publico}" if event and event.token_publico else ''
        my_events_path = '/meus_eventos'

        event_details_url = (
            f"{app_url}{event_path}" if app_url and event_path else
            event_path or
            (f"{app_url}{my_events_path}" if app_url else my_events_path)
        )
        my_events_url = f"{app_url}{my_events_path}" if app_url else my_events_path

        event_date_value = activity.data_atv or (event.data_inicio if event else None)
        event_time_value = activity.hora_atv or (event.hora_inicio if event else None)
        event_date = event_date_value.strftime('%d/%m/%Y') if event_date_value else '-'
        event_time = event_time_value.strftime('%H:%M') if event_time_value else '-'
        event_name = event.nome if event and event.nome else activity.nome

        self.notification_service.send_email_task(
            to_email=user.email,
            subject=f"Você foi adicionado ao evento: {event_name}",
            template_name='manual_enrollment_confirmation.html',
            template_data={
                'user_name': user.nome or user.username,
                'event_name': event_name,
                'activity_name': activity.nome,
                'event_date': event_date,
                'event_time': event_time,
                'event_location': activity.local or '-',
                'event_description': activity.descricao or (event.descricao if event else ''),
                'event_details_url': event_details_url,
                'my_events_url': my_events_url,
                'year': datetime.now().year,
            },
        )

    def parse_users_csv(self, file_stream):
        """Parses CSV rows for user import and returns normalized row records."""
        try:
            content = file_stream.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            return {
                'ok': False,
                'message': 'Falha ao ler CSV.',
                'rows': [],
                'errors': ['Salve o arquivo CSV em UTF-8 e tente novamente.'],
                'ignored_columns': [],
            }

        stream = io.StringIO(content, newline=None)
        csv_input = csv.DictReader(stream)
        if not csv_input.fieldnames:
            return {
                'ok': False,
                'message': 'CSV inválido.',
                'rows': [],
                'errors': ['Cabeçalho não encontrado no CSV.'],
                'ignored_columns': [],
            }

        csv_input.fieldnames = [self._normalize_import_header(field) for field in csv_input.fieldnames]
        required_headers = {'nome', 'cpf'}
        missing_headers = sorted(required_headers - set(csv_input.fieldnames))
        if missing_headers:
            return {
                'ok': False,
                'message': 'Cabeçalho inválido.',
                'rows': [],
                'errors': [f"Colunas obrigatórias ausentes: {', '.join(missing_headers)}"],
                'ignored_columns': [],
            }

        supported_headers = {'nome', 'email', 'cpf', 'ra', 'role', 'curso', 'can_create_events', 'password'}
        ignored_columns = [field for field in csv_input.fieldnames if field not in supported_headers]
        present_fields = {field: field in csv_input.fieldnames for field in supported_headers}

        parsed_rows = []
        for row_number, row in enumerate(csv_input, start=2):
            normalized_row = {
                self._normalize_import_header(key): (value or '').strip()
                for key, value in row.items()
                if key is not None
            }
            if not any(normalized_row.values()):
                continue

            parsed_rows.append({
                'row_number': row_number,
                'nome': normalized_row.get('nome', ''),
                'email': normalized_row.get('email', ''),
                'cpf_raw': normalized_row.get('cpf', ''),
                'ra': normalized_row.get('ra', ''),
                'role_raw': normalized_row.get('role', ''),
                'curso_nome': normalized_row.get('curso', ''),
                'can_create_events_raw': normalized_row.get('can_create_events', ''),
                'password': normalized_row.get('password', ''),
                'has_email': present_fields['email'],
                'has_ra': present_fields['ra'],
                'has_role': present_fields['role'],
                'has_course': present_fields['curso'],
                'has_can_create_events': present_fields['can_create_events'],
                'has_password': present_fields['password'],
            })

        return {
            'ok': True,
            'message': 'Arquivo lido com sucesso.',
            'rows': parsed_rows,
            'errors': [],
            'ignored_columns': ignored_columns,
        }

    def process_user_csv_record(self, row_data):
        """Processes one CSV user row and commits per-row for real-time progress updates."""
        row_number = row_data.get('row_number')
        nome = (row_data.get('nome') or '').strip()
        email = (row_data.get('email') or '').strip().lower() or None
        cpf_raw = row_data.get('cpf_raw', '')
        ra = (row_data.get('ra') or '').strip() or None
        role_raw = row_data.get('role_raw', '')
        role = self._normalize_import_role(role_raw) if str(role_raw or '').strip() else None
        curso_nome = (row_data.get('curso_nome') or '').strip()
        can_create_events = self._coerce_bool(row_data.get('can_create_events_raw'))
        password = row_data.get('password') or ''

        has_email = bool(row_data.get('has_email'))
        has_ra = bool(row_data.get('has_ra'))
        has_role = bool(row_data.get('has_role'))
        has_course = bool(row_data.get('has_course'))
        has_can_create_events = bool(row_data.get('has_can_create_events'))
        has_password = bool(row_data.get('has_password'))

        cpf_digits = self._normalize_cpf_digits(cpf_raw)
        cpf_masked = self._format_cpf_mask(cpf_digits)

        base_payload = {
            'row_number': row_number,
            'nome': nome,
            'cpf': cpf_masked or cpf_raw,
            'curso': curso_nome,
            'ra': ra,
        }

        if len(cpf_digits) != 11 or not cpf_masked:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': f"CPF inválido ({cpf_raw}).",
            }

        if not nome:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': 'NOME ausente.',
            }

        allowed_roles = {'admin', 'professor', 'coordenador', 'gestor', 'extensao', 'participante'}
        if has_role and role_raw and role not in allowed_roles:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': f"Perfil inválido ({role_raw}).",
            }

        course_obj = None
        if has_course and curso_nome:
            course_obj = Course.query.filter(Course.nome.ilike(curso_nome)).first()
            if not course_obj:
                db.session.rollback()
                return {
                    **base_payload,
                    'status': 'error',
                    'message': f"Curso não encontrado ({curso_nome}).",
                }

        existing = self._find_user_by_cpf_flexible(cpf_masked)
        existing_username = existing.username if existing else None

        if email:
            email_owner = User.query.filter(func.lower(User.email) == email).first()
            if email_owner and email_owner.username != existing_username:
                db.session.rollback()
                return {
                    **base_payload,
                    'status': 'error',
                    'message': f"E-mail {email} já pertence a outro usuário.",
                }

        if ra:
            ra_owner = User.query.filter_by(ra=ra).first()
            if ra_owner and ra_owner.username != existing_username:
                db.session.rollback()
                return {
                    **base_payload,
                    'status': 'error',
                    'message': f"RA {ra} já pertence a outro usuário.",
                }

        if existing:
            changed = False

            if existing.nome != nome:
                existing.nome = nome
                changed = True

            if has_email and email and (existing.email or '').strip().lower() != email:
                existing.email = email
                changed = True

            if has_ra and ra and existing.ra != ra:
                existing.ra = ra
                changed = True

            if has_role and role and existing.role != role:
                existing.role = role
                changed = True

            if has_course and curso_nome and course_obj and existing.course_id != course_obj.id:
                existing.course_id = course_obj.id
                changed = True

            if has_can_create_events and existing.can_create_events != can_create_events:
                existing.can_create_events = can_create_events
                changed = True

            if has_password and password:
                existing.set_password(password)
                changed = True

            if changed:
                db.session.commit()
                return {
                    **base_payload,
                    'status': 'updated',
                    'message': f"Usuário {existing.username} atualizado.",
                }

            db.session.rollback()
            return {
                **base_payload,
                'status': 'unchanged',
                'message': f"Usuário {existing.username} sem alteração.",
            }

        user = User(
            username=cpf_digits,
            email=email,
            nome=nome,
            cpf=cpf_masked,
            ra=ra,
            role=role or 'participante',
            course_id=course_obj.id if course_obj else None,
            can_create_events=can_create_events if has_can_create_events else False,
        )
        user.set_password(password or cpf_digits)
        db.session.add(user)
        db.session.commit()
        return {
            **base_payload,
            'status': 'created',
            'message': f"Usuário {user.username} criado.",
        }

    def import_users_csv(self, file_stream):
        """Imports and upserts users from CSV using CPF as the stable key."""
        parsed = self.parse_users_csv(file_stream)
        if not parsed.get('ok'):
            return {
                'message': parsed.get('message', 'Falha ao ler CSV.'),
                'total_rows': 0,
                'processed_rows': 0,
                'created': 0,
                'updated': 0,
                'unchanged': 0,
                'errors': parsed.get('errors', []),
                'ignored_columns': parsed.get('ignored_columns', []),
            }

        created = 0
        updated = 0
        unchanged = 0
        errors = []
        processed_rows = len(parsed.get('rows', []))

        for row_data in parsed.get('rows', []):
            outcome = self.process_user_csv_record(row_data)
            status = outcome.get('status')
            if status == 'created':
                created += 1
            elif status == 'updated':
                updated += 1
            elif status == 'unchanged':
                unchanged += 1
            else:
                errors.append(f"Linha {outcome.get('row_number')}: {outcome.get('message')}")

        return {
            'message': f"Processadas {processed_rows} linhas.",
            'total_rows': processed_rows,
            'processed_rows': processed_rows,
            'created': created,
            'updated': updated,
            'unchanged': unchanged,
            'errors': errors,
            'ignored_columns': parsed.get('ignored_columns', []),
        }

    @staticmethod
    def _normalize_cpf_digits(cpf):
        if not cpf:
            return ''
        return ''.join(ch for ch in str(cpf) if ch.isdigit())

    @staticmethod
    def _format_cpf_mask(cpf_digits):
        if len(cpf_digits) != 11:
            return None
        return f"{cpf_digits[0:3]}.{cpf_digits[3:6]}.{cpf_digits[6:9]}-{cpf_digits[9:11]}"

    def _find_user_by_cpf_flexible(self, cpf_raw):
        cpf_digits = self._normalize_cpf_digits(cpf_raw)
        if len(cpf_digits) != 11:
            return None

        cpf_masked = self._format_cpf_mask(cpf_digits)
        candidates = [cpf_raw, cpf_digits, cpf_masked]
        seen = set()

        for candidate in candidates:
            if not candidate:
                continue
            normalized = str(candidate).strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            user = self.user_repo.get_by_cpf(normalized)
            if user:
                return user

        return None

    def parse_students_xlsx(self, file_stream):
        """Parses XLSX rows for student import and returns normalized row records."""
        workbook = load_workbook(filename=file_stream, read_only=True, data_only=True)
        sheet = workbook.active

        expected_headers = [
            'ALUNO_NOME', 'IES', 'CURSO', 'TURMA', 'CPF', 'DATANASCIMENTO',
            'SEXO', 'ESTADOCIVIL', 'MAE', 'NIVEL ESCOLAR', 'RA', 'TURNO',
            'PERIODO', 'RUA_NUMERO', 'BAIRRO', 'CEP', 'MUNICIPIO', 'ESTADO',
            'Total Geral'
        ]
        required_headers = {'ALUNO_NOME', 'CURSO', 'CPF'}

        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        if not rows:
            return {
                'ok': False,
                'message': 'Arquivo XLSX vazio.',
                'rows': [],
                'errors': ['Arquivo sem linhas para processamento.'],
                'ignored_columns': expected_headers[1:]
            }

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        header_map = {h.upper(): idx for idx, h in enumerate(headers) if h}

        missing_required = sorted([h for h in required_headers if h not in header_map])
        if missing_required:
            return {
                'ok': False,
                'message': 'Cabeçalho inválido.',
                'rows': [],
                'errors': [f"Colunas obrigatórias ausentes: {', '.join(missing_required)}"],
                'ignored_columns': []
            }

        ignored_columns = [col for col in expected_headers if col not in {'ALUNO_NOME', 'CURSO', 'CPF', 'RA'}]

        idx_nome = header_map.get('ALUNO_NOME')
        idx_curso = header_map.get('CURSO')
        idx_cpf = header_map.get('CPF')
        idx_ra = header_map.get('RA')
        idx_email = header_map.get('EMAIL')

        parsed_rows = []
        for row_number, row in enumerate(rows[1:], start=2):
            if row is None:
                continue

            nome = str(row[idx_nome]).strip() if idx_nome is not None and idx_nome < len(row) and row[idx_nome] is not None else ''
            curso_nome = str(row[idx_curso]).strip() if idx_curso is not None and idx_curso < len(row) and row[idx_curso] is not None else ''
            cpf_raw = str(row[idx_cpf]).strip() if idx_cpf is not None and idx_cpf < len(row) and row[idx_cpf] is not None else ''
            ra = str(row[idx_ra]).strip() if idx_ra is not None and idx_ra < len(row) and row[idx_ra] is not None else None
            email = str(row[idx_email]).strip() if idx_email is not None and idx_email < len(row) and row[idx_email] is not None else ''

            if not any([nome, curso_nome, cpf_raw, ra, email]):
                continue

            parsed_rows.append({
                'row_number': row_number,
                'nome': nome,
                'curso_nome': curso_nome,
                'cpf_raw': cpf_raw,
                'ra': ra,
                'email': email,
            })

        return {
            'ok': True,
            'message': 'Arquivo lido com sucesso.',
            'rows': parsed_rows,
            'errors': [],
            'ignored_columns': ignored_columns,
        }

    def process_student_record(self, row_data):
        """Processes one student row and commits per-row for real-time progress updates."""
        row_number = row_data.get('row_number')
        nome = row_data.get('nome', '')
        curso_nome = row_data.get('curso_nome', '')
        cpf_raw = row_data.get('cpf_raw', '')
        ra = row_data.get('ra')
        email = row_data.get('email', '')

        cpf_digits = self._normalize_cpf_digits(cpf_raw)
        cpf_masked = self._format_cpf_mask(cpf_digits)

        base_payload = {
            'row_number': row_number,
            'nome': nome,
            'cpf': cpf_masked or cpf_raw,
            'curso': curso_nome,
            'ra': ra,
        }

        if len(cpf_digits) != 11 or not cpf_masked:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': f"CPF inválido ({cpf_raw}).",
            }

        if not nome:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': 'ALUNO_NOME ausente.',
            }

        if not curso_nome:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': 'CURSO ausente.',
            }

        course_obj = Course.query.filter(Course.nome.ilike(curso_nome)).first()
        if not course_obj:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': f"Curso não encontrado ({curso_nome}).",
            }

        existing = self._find_user_by_cpf_flexible(cpf_masked)
        if existing:
            if ra:
                ra_owner = User.query.filter(User.ra == ra, User.username != existing.username).first()
                if ra_owner:
                    db.session.rollback()
                    return {
                        **base_payload,
                        'status': 'error',
                        'message': f"RA {ra} já pertence a outro usuário.",
                    }

            changed = False
            if existing.nome != nome:
                existing.nome = nome
                changed = True
            if existing.course_id != course_obj.id:
                existing.course_id = course_obj.id
                changed = True
            if ra and existing.ra != ra:
                existing.ra = ra
                changed = True

            if changed:
                db.session.commit()
                return {
                    **base_payload,
                    'status': 'updated',
                    'message': f"Usuário {existing.username} atualizado.",
                }

            db.session.rollback()
            return {
                **base_payload,
                'status': 'unchanged',
                'message': f"Usuário {existing.username} sem alteração.",
            }

        if not email:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': 'Novo aluno sem EMAIL, criação bloqueada.',
            }

        username = cpf_digits
        username_owner = self.user_repo.get_by_username(username)
        if username_owner:
            db.session.rollback()
            return {
                **base_payload,
                'status': 'error',
                'message': f"Username {username} já existe para outro usuário.",
            }

        if ra:
            ra_owner = User.query.filter_by(ra=ra).first()
            if ra_owner:
                db.session.rollback()
                return {
                    **base_payload,
                    'status': 'error',
                    'message': f"RA {ra} já cadastrado.",
                }

        user = User(
            username=username,
            email=email,
            nome=nome,
            cpf=cpf_masked,
            ra=ra,
            role='participante',
            course_id=course_obj.id,
            can_create_events=False,
        )
        user.set_password(cpf_digits)
        db.session.add(user)
        db.session.commit()
        return {
            **base_payload,
            'status': 'created',
            'message': f"Usuário {username} criado.",
        }

    def import_students_xlsx(self, file_stream):
        """
        Imports and upserts students from an academic XLSX file.

        Rules:
        - Upsert key: CPF
        - Existing users: update only nome, curso and ra
        - New users: create only if email is present in the row
        - New users always receive role='participante'
        - Username for new users is CPF digits (no punctuation)
        """
        parsed = self.parse_students_xlsx(file_stream)
        if not parsed.get('ok'):
            return {
                'message': parsed.get('message', 'Falha ao ler XLSX.'),
                'total_rows': 0,
                'created': 0,
                'updated': 0,
                'unchanged': 0,
                'errors': parsed.get('errors', []),
                'ignored_columns': parsed.get('ignored_columns', []),
            }

        created = 0
        updated = 0
        unchanged = 0
        errors = []
        processed_rows = len(parsed.get('rows', []))

        for row_data in parsed.get('rows', []):
            outcome = self.process_student_record(row_data)
            status = outcome.get('status')
            if status == 'created':
                created += 1
            elif status == 'updated':
                updated += 1
            elif status == 'unchanged':
                unchanged += 1
            else:
                errors.append(f"Linha {outcome.get('row_number')}: {outcome.get('message')}")

        return {
            'message': f"Processadas {processed_rows} linhas.",
            'total_rows': processed_rows,
            'created': created,
            'updated': updated,
            'unchanged': unchanged,
            'errors': errors,
            'ignored_columns': parsed.get('ignored_columns', []),
        }

    def update_user_permissions(self, username, can_create_events):
        """Updates specific permission flags for a user."""
        user = self.user_repo.get_by_username(username)
        if not user:
            return False, "Usuário não encontrado."
            
        user.can_create_events = can_create_events
        self.user_repo.update()
        return True, "Permissões atualizadas."

    def bulk_update_permissions_by_course(self, course_id, can_create_events, role_filter='professor'):
        """
        Updates permissions for all users of a specific course and role.
        """
        users = User.query.filter_by(course_id=course_id, role=role_filter).all()
        count = 0
        for user in users:
            user.can_create_events = can_create_events
            count += 1
        
        db.session.commit()
        return count, f"{count} professores atualizados."
